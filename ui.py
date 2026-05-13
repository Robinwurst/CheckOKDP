# ui.py
"""
Главное окно приложения.
Включает:
- выбор файлов
- кнопку «Обработать» (выполняет чтение, парсинг и объединение)
- таблицу с результатами (QTableView + QStandardItemModel)
- кнопку «Сохранить результат»
"""

from openpyxl.styles import PatternFill
import pandas as pd
from pathlib import Path
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QLabel,
    QLineEdit,
    QFileDialog,
    QMessageBox,
    QTableView,
    QStatusBar,
    QGroupBox,
    QGridLayout,
    QSizePolicy,
    QFrame,
    QStyle,
    QToolButton,
    QAbstractItemView,
    QComboBox,
)
from PySide6.QtCore import QModelIndex, QSize, Qt, QSortFilterProxyModel
from PySide6.QtGui import QStandardItemModel, QStandardItem, QColor, QBrush
from services import read_registry_file, read_code_name_file, analyze_row


class _DiffRowFilterProxy(QSortFilterProxyModel):
    """Фильтр строк по флагу «отличается от исходника»; корректно работает при сортировке."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._diff: list[bool] = []
        self._mode = "all"

    def set_diff_rows(self, flags: list[bool]) -> None:
        self._diff = list(flags)
        self.invalidateFilter()

    def set_filter_mode(self, mode: object) -> None:
        m = "all"
        if isinstance(mode, str) and mode in ("all", "diff", "same"):
            m = mode
        elif mode is not None:
            ms = str(mode)
            if ms in ("all", "diff", "same"):
                m = ms
        self._mode = m
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        if self._mode == "all":
            return True
        if source_parent.isValid():
            return True
        if source_row < 0 or source_row >= len(self._diff):
            return True
        differs = self._diff[source_row]
        if self._mode == "diff":
            return differs
        return not differs


_APP_STYLESHEET = """
/* Светлая тема: карточки, градиент фона, акценты */
QWidget {
    font-family: "Segoe UI", "SF Pro Text", "Roboto", "Helvetica Neue", sans-serif;
    font-size: 13px;
    color: #1c2430;
}
QMainWindow, QWidget#CentralRoot {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #e8edf5, stop:0.45 #eef2f8, stop:1 #f6f8fc);
}
QFrame#HeroCard {
    background-color: #ffffff;
    border: 1px solid #dde4ee;
    border-radius: 16px;
}
QFrame#HeroAccent {
    border: none;
    border-radius: 4px;
    min-height: 4px;
    max-height: 4px;
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #2563eb, stop:0.55 #4f46e5, stop:1 #7c3aed);
}
QLabel#AppKicker {
    color: #3b5bdb;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.12em;
}
QLabel#AppTitle {
    font-size: 22px;
    font-weight: 700;
    color: #0f172a;
    letter-spacing: -0.03em;
}
QLabel#AppHint {
    color: #5b6475;
    font-size: 12px;
}
QLabel#FieldLabel {
    color: #64748b;
    font-size: 11px;
    font-weight: 600;
}
QLabel#SectionTitle {
    color: #334155;
    font-size: 12px;
    font-weight: 700;
    letter-spacing: 0.04em;
    margin-top: 4px;
}
QFrame#ActionCard {
    background-color: #ffffff;
    border: 1px solid #dde4ee;
    border-radius: 14px;
}
QFrame#TableCard {
    background-color: #ffffff;
    border: 1px solid #dde4ee;
    border-radius: 14px;
}
QComboBox#ResultFilter {
    min-height: 28px;
    padding: 4px 12px 4px 10px;
    border: 1px solid #c9d1e0;
    border-radius: 8px;
    background-color: #ffffff;
    color: #1e293b;
    font-weight: 600;
    font-size: 12px;
}
QComboBox#ResultFilter:hover {
    border-color: #94a3b8;
}
QComboBox#ResultFilter::drop-down {
    border: none;
    width: 22px;
}
QGroupBox {
    font-weight: 600;
    color: #111827;
    border: 1px solid #dde4ee;
    border-radius: 14px;
    margin-top: 18px;
    padding: 16px 14px 14px 14px;
    background-color: #ffffff;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 16px;
    padding: 0 8px;
    color: #475569;
    font-size: 12px;
    letter-spacing: 0.03em;
}
QLineEdit {
    min-height: 26px;
    padding: 6px 11px;
    border: 1px solid #c9d1e0;
    border-radius: 9px;
    background-color: #fafbfd;
    color: #1c2430;
    selection-background-color: #3b6cff;
    selection-color: #ffffff;
}
QLineEdit:focus {
    border: 1px solid #2563eb;
    background-color: #ffffff;
}
QPushButton#PrimaryButton {
    min-height: 32px;
    padding: 8px 20px;
    border-radius: 9px;
    border: none;
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #3b82f6, stop:1 #2563eb);
    color: #ffffff;
    font-weight: 600;
}
QPushButton#PrimaryButton:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #4f8ff8, stop:1 #2f6feb);
}
QPushButton#PrimaryButton:pressed {
    background: #1d4ed8;
}
QPushButton#PrimaryButton:disabled {
    background: #c5d4f7;
    color: #64748b;
}
QPushButton#SecondaryButton {
    min-height: 32px;
    padding: 8px 20px;
    border-radius: 9px;
    border: 1px solid #c9d1e0;
    background-color: #ffffff;
    color: #1e293b;
    font-weight: 600;
}
QPushButton#SecondaryButton:hover {
    background-color: #f8fafc;
    border-color: #94a3b8;
}
QPushButton#SecondaryButton:pressed {
    background-color: #f1f5f9;
}
QPushButton#SecondaryButton:disabled {
    color: #9aa4b2;
    border-color: #dde3ee;
    background-color: #f8fafc;
}
QPushButton#TonalButton {
    min-height: 32px;
    padding: 8px 14px;
    border-radius: 9px;
    border: 1px solid #c9d1e0;
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #ffffff, stop:1 #f1f5f9);
    color: #1e293b;
    font-weight: 600;
}
QPushButton#TonalButton:hover {
    background: #f8fafc;
    border-color: #94a3b8;
}
QPushButton#TonalButton:pressed {
    background: #e2e8f0;
}
QTableView {
    border: none;
    border-radius: 12px;
    gridline-color: transparent;
    background-color: #ffffff;
    alternate-background-color: #f8fafc;
    color: #1c2430;
    selection-background-color: #bfdbfe;
    selection-color: #0f172a;
}
QTableView::item {
    padding: 8px 10px;
}
QHeaderView::section {
    padding: 10px 12px;
    border: none;
    border-bottom: 2px solid #e2e8f0;
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #ffffff, stop:1 #f8fafc);
    color: #334155;
    font-weight: 700;
    font-size: 12px;
}
QStatusBar {
    background-color: #ffffff;
    color: #64748b;
    border-top: 1px solid #e2e8f0;
    padding: 2px 0;
}
QToolButton#AuthorChip {
    border: 1px solid #e2e8f0;
    border-radius: 16px;
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #ffffff, stop:1 #f1f5f9);
    padding: 0;
    margin: 3px 4px 3px 10px;
}
QToolButton#AuthorChip:hover {
    border-color: #cbd5e1;
    background: #f8fafc;
}
QToolButton#AuthorChip:pressed {
    background: #e2e8f0;
}
QFrame#AuthorBadge {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #ffffff, stop:1 #f1f5f9);
    border: 1px solid #e2e8f0;
    border-radius: 999px;
    margin: 3px 10px 3px 0;
    padding: 0;
}
QLabel#AuthorMicro {
    color: #94a3b8;
    font-size: 9px;
    font-weight: 700;
    letter-spacing: 0.14em;
}
QLabel#AuthorHandle {
    color: #334155;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.02em;
}
QLabel#AuthorName {
    color: #64748b;
    font-size: 10px;
    font-weight: 500;
}
QScrollBar:vertical {
    width: 11px;
    margin: 0;
    background: #f8fafc;
    border-radius: 6px;
}
QScrollBar::handle:vertical {
    min-height: 32px;
    background: #cbd5e1;
    border-radius: 5px;
    margin: 2px;
}
QScrollBar::handle:vertical:hover {
    background: #94a3b8;
}
QScrollBar:horizontal {
    height: 11px;
    margin: 0;
    background: #f8fafc;
    border-radius: 6px;
}
QScrollBar::handle:horizontal {
    min-width: 32px;
    background: #cbd5e1;
    border-radius: 5px;
    margin: 2px;
}
QMessageBox {
    background-color: #ffffff;
}
QMessageBox QLabel {
    color: #1c2430;
}
"""


def apply_app_theme(app: QApplication) -> None:
    """Fusion + светлый QSS: предсказуемый вид на Windows (exe) и на macOS."""
    app.setStyle("Fusion")
    app.setStyleSheet(_APP_STYLESHEET)


class MainWindow(QMainWindow):
    """Главное окно приложения."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Объединение по ОКПД2")
        self.setMinimumSize(880, 560)
        self.resize(1024, 700)

        # Путь к файлам и DataFrames
        self.registry_path: str | None = None
        self.codename_path: str | None = None
        self.registry_df: pd.DataFrame | None = None
        self.codename_df: pd.DataFrame | None = None
        self._row_highlight: list[bool] = []
        self._table_proxy = _DiffRowFilterProxy(self)

        self._setup_ui()

    # ---------------------------------------------------------------------
    # UI
    # ---------------------------------------------------------------------
    def _setup_ui(self):
        central = QWidget()
        central.setObjectName("CentralRoot")
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(22, 22, 22, 18)
        main_layout.setSpacing(18)

        hero = QFrame()
        hero.setObjectName("HeroCard")
        hero.setFrameShape(QFrame.Shape.NoFrame)
        hero.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum)
        hero_layout = QVBoxLayout(hero)
        hero_layout.setContentsMargins(22, 20, 22, 20)
        hero_layout.setSpacing(12)

        accent = QFrame()
        accent.setObjectName("HeroAccent")
        accent.setFrameShape(QFrame.Shape.NoFrame)
        accent.setFixedHeight(4)
        hero_layout.addWidget(accent)

        kicker = QLabel("ОКПД2  ·  слияние с реестром")
        kicker.setObjectName("AppKicker")
        hero_layout.addWidget(kicker)

        title = QLabel("Сопоставление кодов из файла с реестром ОКПД2 (поиск самого длинного кода)")
        title.setObjectName("AppTitle")
        title.setWordWrap(True)
        hero_layout.addWidget(title)

        hint = QLabel(
            "Реестр — Excel с колонками «Код» и «Название» (шапка с 7-й строки; лист с данными "
            "определяется автоматически, имя листа не важно). Скачивать: https://classifikators.ru/okpd "
            "Второй файл — одна колонка: «код название» через пробел; лист с данными "
            "ищется автоматически. Поддерживаются .xlsx, .xls, .xlsm, .xlsb."
        )
        hint.setObjectName("AppHint")
        hint.setWordWrap(True)
        hero_layout.addWidget(hint)

        main_layout.addWidget(hero)

        files_group = QGroupBox("Файлы для обработки")
        files_grid = QGridLayout(files_group)
        files_grid.setColumnStretch(1, 1)
        files_grid.setHorizontalSpacing(12)
        files_grid.setVerticalSpacing(12)

        self.registry_line = QLineEdit()
        self.registry_line.setPlaceholderText("Выберите Excel реестра…")
        self.registry_line.setReadOnly(True)
        self.registry_line.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.registry_browse = QPushButton("Обзор…")
        self.registry_browse.setObjectName("TonalButton")
        self.registry_browse.setFixedWidth(128)
        self.registry_browse.setCursor(Qt.CursorShape.PointingHandCursor)

        self.codename_line = QLineEdit()
        self.codename_line.setPlaceholderText("Выберите Excel «код + название»…")
        self.codename_line.setReadOnly(True)
        self.codename_line.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.codename_browse = QPushButton("Обзор…")
        self.codename_browse.setObjectName("TonalButton")
        self.codename_browse.setFixedWidth(128)
        self.codename_browse.setCursor(Qt.CursorShape.PointingHandCursor)

        lbl_registry = QLabel("Реестр ОКПД2")
        lbl_registry.setObjectName("FieldLabel")
        lbl_codes = QLabel("Код и название")
        lbl_codes.setObjectName("FieldLabel")
        files_grid.addWidget(lbl_registry, 0, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        files_grid.addWidget(self.registry_line, 0, 1)
        files_grid.addWidget(self.registry_browse, 0, 2)
        files_grid.addWidget(lbl_codes, 1, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        files_grid.addWidget(self.codename_line, 1, 1)
        files_grid.addWidget(self.codename_browse, 1, 2)
        main_layout.addWidget(files_group)

        action_card = QFrame()
        action_card.setObjectName("ActionCard")
        action_card.setFrameShape(QFrame.Shape.NoFrame)
        action_card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum)
        action_layout = QHBoxLayout(action_card)
        action_layout.setContentsMargins(16, 14, 16, 14)
        action_layout.setSpacing(12)

        self.process_btn = QPushButton("Обработать")
        self.process_btn.setObjectName("PrimaryButton")
        self.process_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.save_btn = QPushButton("Сохранить в Excel…")
        self.save_btn.setObjectName("SecondaryButton")
        self.save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.save_btn.setEnabled(False)
        action_layout.addWidget(self.process_btn)
        action_layout.addWidget(self.save_btn)
        action_layout.addStretch(1)
        main_layout.addWidget(action_card)

        section_row = QWidget()
        section_layout = QHBoxLayout(section_row)
        section_layout.setContentsMargins(0, 0, 0, 0)
        section_layout.setSpacing(12)
        section_title = QLabel("Результат")
        section_title.setObjectName("SectionTitle")
        section_layout.addWidget(section_title)
        section_layout.addStretch(1)
        filter_lbl = QLabel("Показать:")
        filter_lbl.setObjectName("FieldLabel")
        section_layout.addWidget(filter_lbl, alignment=Qt.AlignmentFlag.AlignVCenter)
        self._result_filter = QComboBox()
        self._result_filter.setObjectName("ResultFilter")
        self._result_filter.setMinimumWidth(260)
        self._result_filter.setEnabled(False)
        self._result_filter.addItem("Все строки", "all")
        self._result_filter.addItem("Только подсвеченные", "diff")
        self._result_filter.addItem("Только без подсветки", "same")
        self._result_filter.currentIndexChanged.connect(self._on_result_filter_changed)
        section_layout.addWidget(self._result_filter, alignment=Qt.AlignmentFlag.AlignVCenter)
        main_layout.addWidget(section_row)

        table_shell = QFrame()
        table_shell.setObjectName("TableCard")
        table_shell.setFrameShape(QFrame.Shape.NoFrame)
        table_layout = QVBoxLayout(table_shell)
        table_layout.setContentsMargins(10, 10, 10, 10)

        self.table_view = QTableView()
        self.table_view.setAlternatingRowColors(False)
        self.table_view.setShowGrid(False)
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_view.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table_view.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_view.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_view.setWordWrap(True)
        self.table_view.setTextElideMode(Qt.TextElideMode.ElideNone)
        self.table_view.verticalHeader().setDefaultSectionSize(40)
        self.table_view.verticalHeader().setMinimumSectionSize(28)
        self.table_view.horizontalHeader().setHighlightSections(False)
        self.table_view.horizontalHeader().setSectionsClickable(True)
        self.table_view.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        table_layout.addWidget(self.table_view)

        nav_hint = QLabel(
            "Подсветка (персиковый): код не найден в реестре; найдено несколько вариантов "
            "с одинаковой «глубиной» кода; один вариант, но код или наименование в файле "
            "не совпадают с реестром. Фильтр: все / только подсвеченные / без подсветки. "
            "При сохранении в Excel подсветка совпадает с таблицей. "
            "Колёсико — прокрутка; заголовок — сортировка; Shift/Ctrl — несколько строк."
        )
        nav_hint.setObjectName("AppHint")
        nav_hint.setWordWrap(True)
        table_layout.addWidget(nav_hint)

        main_layout.addWidget(table_shell, stretch=1)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.addPermanentWidget(self._author_badge())

        self.registry_browse.clicked.connect(self._browse_registry)
        self.codename_browse.clicked.connect(self._browse_codename)
        self.process_btn.clicked.connect(self._process_files)
        self.save_btn.clicked.connect(self._save_result)

        self._apply_standard_icons()

    def _apply_standard_icons(self) -> None:
        style = self.style()
        if style is None:
            return
        icon_sz = QSize(20, 20)
        open_icon = style.standardIcon(QStyle.StandardPixmap.SP_DialogOpenButton)
        apply_icon = style.standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton)
        save_icon = style.standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton)

        self.registry_browse.setIcon(open_icon)
        self.registry_browse.setIconSize(icon_sz)
        self.codename_browse.setIcon(open_icon)
        self.codename_browse.setIconSize(icon_sz)

        self.process_btn.setIcon(apply_icon)
        self.process_btn.setIconSize(icon_sz)
        self.save_btn.setIcon(save_icon)
        self.save_btn.setIconSize(icon_sz)

    def _author_badge(self) -> QWidget:
        root = QWidget()
        root.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Maximum)
        row = QHBoxLayout(root)
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(8)

        chip = QToolButton()
        chip.setObjectName("AuthorChip")
        chip.setAutoRaise(True)
        chip.setCursor(Qt.CursorShape.PointingHandCursor)
        chip.setFixedSize(32, 32)
        chip.setToolTip(
            "RobinWurst\n"
            "Роберт Ерузалимский\n\n"
            "Нажмите — показать или скрыть подпись."
        )
        chip.setToolTipDuration(8000)
        style = self.style()
        if style is not None:
            chip.setIcon(style.standardIcon(QStyle.StandardPixmap.SP_MessageBoxInformation))
            chip.setIconSize(QSize(18, 18))

        detail = QFrame()
        detail.setObjectName("AuthorBadge")
        detail.setFrameShape(QFrame.Shape.NoFrame)
        detail.setVisible(False)
        detail.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Maximum)

        outer = QHBoxLayout(detail)
        outer.setContentsMargins(12, 5, 14, 5)
        outer.setSpacing(10)

        micro = QLabel("АВТОР")
        micro.setObjectName("AuthorMicro")

        text_col = QVBoxLayout()
        text_col.setSpacing(1)
        text_col.setContentsMargins(0, 0, 0, 0)
        handle = QLabel("RobinWurst")
        handle.setObjectName("AuthorHandle")
        name = QLabel("Роберт Ерузалимский")
        name.setObjectName("AuthorName")
        text_col.addWidget(handle)
        text_col.addWidget(name)

        outer.addWidget(micro, alignment=Qt.AlignmentFlag.AlignVCenter)
        outer.addLayout(text_col)

        def toggle_detail() -> None:
            detail.setVisible(not detail.isVisible())

        chip.clicked.connect(toggle_detail)

        row.addWidget(chip, alignment=Qt.AlignmentFlag.AlignVCenter)
        row.addWidget(detail, alignment=Qt.AlignmentFlag.AlignVCenter)

        return root

    # ---------------------------------------------------------------------
    # Файлы
    # ---------------------------------------------------------------------
    def _browse_registry(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл реестра", "", "Excel (*.xlsx *.xls *.xlsm *.xlsb)"
        )
        if path:
            self.registry_path = path
            self.registry_line.setText(path)

    def _browse_codename(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл код+название", "", "Excel (*.xlsx *.xls *.xlsm *.xlsb)"
        )
        if path:
            self.codename_path = path
            self.codename_line.setText(path)

    # ---------------------------------------------------------------------
    # Обработка файлов
    # ---------------------------------------------------------------------
    def _process_files(self):
        if not self.registry_path or not self.codename_path:
            QMessageBox.warning(self, "Внимание", "Пожалуйста, выберите оба файла.")
            return

        self.status_bar.showMessage("Чтение файлов…", 2000)
        try:
            self.registry_df = read_registry_file(self.registry_path)
            registry_dict: dict[str, str] = {
                str(row["Код"]).strip(): str(row["Название"]).strip()
                for _, row in self.registry_df.iterrows()
            }

            self.codename_df = read_code_name_file(self.codename_path)

            new_col: list[str] = []
            highlight: list[bool] = []
            for cell in self.codename_df["Value"]:
                text, _in_registry, hi = analyze_row(cell, registry_dict)
                new_col.append(text)
                highlight.append(hi)
            self.codename_df["Объединенный"] = new_col
            self._row_highlight = highlight

            self._show_result()
            self.save_btn.setEnabled(True)

        except Exception as exc:
            self._row_highlight = []
            self._table_proxy.setSourceModel(None)
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось обработать файлы:\n{exc}"
            )
            self.status_bar.clearMessage()

    # ---------------------------------------------------------------------
    # Отображение результата
    # ---------------------------------------------------------------------
    def _show_result(self):
        if self.codename_df is None or self.codename_df.empty:
            QMessageBox.information(self, "Результат", "Нет данных для отображения.")
            return

        rows, cols = self.codename_df.shape
        src_flags = self._row_highlight
        use_row_colors = len(src_flags) == rows
        if use_row_colors:
            diff_rows = list(src_flags)
        else:
            diff_rows = [False] * rows

        brush_same = QBrush(QColor(255, 255, 255))
        brush_diff = QBrush(QColor(255, 243, 224))

        self.table_view.setSortingEnabled(False)
        model = QStandardItemModel(rows, cols)
        model.setHorizontalHeaderLabels(self.codename_df.columns.tolist())

        for row in range(rows):
            for col in range(cols):
                value = self.codename_df.iat[row, col]
                item = QStandardItem(str(value))
                item.setEditable(False)
                if use_row_colors:
                    item.setBackground(brush_diff if diff_rows[row] else brush_same)
                model.setItem(row, col, item)

        self._table_proxy.setSourceModel(model)
        self._table_proxy.set_diff_rows(diff_rows)
        self._table_proxy.set_filter_mode("all")
        self.table_view.setModel(self._table_proxy)
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.verticalHeader().setVisible(False)
        self.table_view.resizeColumnsToContents()
        if rows <= 1200:
            self.table_view.resizeRowsToContents()
        self.table_view.setSortingEnabled(True)

        self._result_filter.setEnabled(True)
        self._result_filter.blockSignals(True)
        self._result_filter.setCurrentIndex(0)
        self._result_filter.blockSignals(False)
        self._apply_table_row_filter()

        n_diff = sum(1 for f in diff_rows if f) if use_row_colors else 0
        if use_row_colors:
            self.status_bar.showMessage(
                f"Строк: {rows}. Требуют внимания (подсветка): {n_diff}, без подсветки: {rows - n_diff}. "
                "Сортировка — по заголовку колонки; фильтр — справа от «Результат».",
                10000,
            )
        else:
            self.status_bar.showMessage(
                f"Строк: {rows}. Сортировка — по заголовку колонки; фильтр — справа от «Результат».",
                8000,
            )

    def _on_result_filter_changed(self, _index: int) -> None:
        self._apply_table_row_filter()

    def _apply_table_row_filter(self) -> None:
        source = self._table_proxy.sourceModel()
        if source is None:
            return
        if len(self._row_highlight) != source.rowCount():
            self._table_proxy.set_filter_mode("all")
            return
        mode = self._result_filter.currentData()
        self._table_proxy.set_filter_mode(mode)

    # ---------------------------------------------------------------------
    # Сохранение
    # ---------------------------------------------------------------------
    def _save_result(self):
        if self.codename_df is None:
            QMessageBox.warning(self, "Внимание", "Нет данных для сохранения.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить результат", "", "Excel файлы (*.xlsx)"
        )
        if not path:
            return

        try:
            path_p = Path(path)
            sheet_name = "Sheet1"
            n_rows_df = len(self.codename_df)
            n_cols = self.codename_df.shape[1]
            hi = self._row_highlight

            with pd.ExcelWriter(path_p, engine="openpyxl") as writer:
                self.codename_df.to_excel(
                    writer, index=False, sheet_name=sheet_name
                )
                if len(hi) == n_rows_df and any(hi):
                    ws = writer.sheets[sheet_name]
                    row_fill = PatternFill(
                        fill_type="solid", fgColor="FFF3E0"
                    )
                    for i, need_hi in enumerate(hi):
                        if not need_hi:
                            continue
                        excel_row = i + 2
                        for c in range(1, n_cols + 1):
                            ws.cell(row=excel_row, column=c).fill = row_fill

            QMessageBox.information(
                self,
                "Готово",
                f"Результат сохранён: {path}\n\n"
                "Подсвеченные в файле строки совпадают с таблицей: нет в реестре, "
                "несколько совпадений, либо расхождение кода/наименования с реестром.",
            )
            self.status_bar.showMessage(f"Сохранено: {path}", 5000)
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось сохранить файл: {exc}"
            )
